from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = Path(r"C:\Users\mudim\Downloads\visa_menged_all_50_countries_estimates_cleaned.xlsx")
COUNTRIES_PATH = ROOT / "data" / "countries_top25.json"
FEE_GUIDES_PATH = ROOT / "data" / "visa_fee_guides.json"
RESEARCH_PATH = ROOT / "data" / "visa_menged_research_seed_may_2026.json"

VISA_TYPE_MAP = {
    "Visitor/Tourist": ("Tourist / Visitor Visa", "tourist-visa", "Tourist / visitor"),
    "Business": ("Business Visa", "business-visa", "Business"),
    "Medical": ("Medical Visa", "medical-visa", "Medical"),
    "Student": ("Student Visa", "student-visa", "Student"),
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    replacements = {
        "\u20ac": "EUR ",
        "\u00a3": "GBP ",
        "\u2013": "-",
        "\u2014": "-",
        "\u00a0": " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    phrase_replacements = {
        "Use as an estimate only; confirm the live official fee before publishing/checkout.": "Use as an estimate only; confirm the live official fee before paying.",
        "verify in official Visa Pricing Estimator before filing": "confirm the final amount on the official Australian fee page before applying",
        "verify estimator": "confirm the final amount on the official Australian fee page",
        "use official pricing estimator before filing": "confirm the final amount on the official Australian fee page before applying",
        "verify in estimator": "confirm the final amount on the official fee page before applying",
    }
    for old, new in phrase_replacements.items():
        text = re.sub(re.escape(old), new, text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def split_urls(value: str) -> list[str]:
    urls = []
    for part in re.split(r"[;\n]+", value):
        part = clean(part)
        if part.startswith("http"):
            urls.append(part)
    return list(dict.fromkeys(urls))


def split_requirements(value: str) -> list[str]:
    value = clean(value)
    if not value:
        return []
    parts = re.split(r"\s*;\s*|\s*\n\s*|(?<=[.])\s+(?=[A-Z])", value)
    requirements = [clean(p).rstrip(".") for p in parts if clean(p)]
    return requirements[:24]


def split_steps(value: str) -> list[str]:
    value = clean(value)
    if not value:
        return []
    # Handles workbook patterns like "1) ... 2) ..." and also semicolon lists.
    if re.search(r"\b1\)", value):
        parts = re.split(r"\s*\d+\)\s*", value)
    else:
        parts = re.split(r"\s*;\s*|\s*\n\s*", value)
    steps = [clean(p).rstrip(".") for p in parts if clean(p)]
    return steps[:12]


def source_entries(country: str, urls: list[str], used_for: str) -> list[dict[str, str]]:
    entries = []
    for index, url in enumerate(urls, start=1):
        entries.append(
            {
                "title": f"{country} official visa source {index}",
                "url": url,
                "used_for": used_for,
            }
        )
    return entries


def merge_sources(existing: list[dict[str, str]], new_sources: list[dict[str, str]]) -> list[dict[str, str]]:
    seen = set()
    merged = []
    for source in [*existing, *new_sources]:
        url = source.get("url", "")
        if not url or url in seen:
            continue
        seen.add(url)
        merged.append(source)
    return merged


def notes_text(*parts: str) -> str:
    kept = [clean(part) for part in parts if clean(part)]
    return " ".join(kept)


def sanitize_value(value: Any) -> Any:
    if isinstance(value, str):
        return clean(value)
    if isinstance(value, list):
        sanitized = [sanitize_value(item) for item in value]
        if all(isinstance(item, str) for item in sanitized):
            return list(dict.fromkeys(item for item in sanitized if item))
        return sanitized
    if isinstance(value, dict):
        return {key: sanitize_value(item) for key, item in value.items()}
    return value


def find_or_create_visa_type(country_obj: dict[str, Any], title: str, slug: str) -> dict[str, Any]:
    visa_types = country_obj.setdefault("visa_types", [])
    for visa_type in visa_types:
        if visa_type.get("visa_type_slug") == slug:
            return visa_type
    visa_type = {
        "visa_type": title,
        "visa_type_slug": slug,
        "summary": f"{title} information for Ethiopian passport holders applying for {country_obj['country']}.",
        "who_can_apply": [
            "Ethiopian passport holders with a valid purpose matching this visa type.",
            "Final eligibility depends on the official portal, embassy or visa center decision.",
        ],
        "requirements": [],
        "documents_needed": [],
        "fees": {},
        "application_steps": [],
        "official_sources": [],
        "data_confidence": "Medium",
    }
    visa_types.append(visa_type)
    return visa_type


def build_document_list(requirements: list[str]) -> list[str]:
    docs = []
    for requirement in requirements:
        if any(
            keyword in requirement.lower()
            for keyword in [
                "passport",
                "photo",
                "form",
                "letter",
                "statement",
                "booking",
                "ticket",
                "invitation",
                "insurance",
                "certificate",
                "admission",
                "acceptance",
                "medical",
                "bank",
                "proof",
                "report",
                "itinerary",
                "registration",
            ]
        ):
            docs.append(requirement)
    return docs[:18] or requirements[:12]


def main() -> None:
    countries = json.loads(COUNTRIES_PATH.read_text(encoding="utf-8"))
    fee_guides = json.loads(FEE_GUIDES_PATH.read_text(encoding="utf-8"))
    research = json.loads(RESEARCH_PATH.read_text(encoding="utf-8"))

    country_slug_by_name = {country["name"].lower(): country["slug"] for country in countries}
    research_by_slug = {country["slug"]: country for country in research}

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Visa_Master"]
    headers = [clean(cell.value) for cell in ws[1]]
    workbook_rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: clean(value) for index, value in enumerate(values)}
        if row.get("Country") and row.get("Visa Type"):
            workbook_rows.append(row)

    evisa_summary: dict[str, dict[str, str]] = {}
    if "eVisa_Summary" in wb.sheetnames:
        ews = wb["eVisa_Summary"]
        eheaders = [clean(cell.value) for cell in ews[1]]
        for values in ews.iter_rows(min_row=2, values_only=True):
            row = {eheaders[index]: clean(value) for index, value in enumerate(values)}
            if row.get("Country"):
                evisa_summary[row["Country"].lower()] = row

    grouped_fee_rows: dict[str, list[dict[str, str]]] = {}

    for row in workbook_rows:
        country_name = row["Country"]
        slug = country_slug_by_name.get(country_name.lower())
        if not slug or slug not in research_by_slug:
            continue

        visa_title, visa_slug, fee_type_label = VISA_TYPE_MAP[row["Visa Type"]]
        country_obj = research_by_slug[slug]
        country_obj["last_verified_approx"] = "May 2026"

        where_to_apply = row["Where to apply"]
        if where_to_apply:
            country_obj["application_channel_from_ethiopia"] = where_to_apply

        urls = split_urls(row["Official source URLs"])
        source_note = notes_text(
            f"Workbook confidence: {row['Confidence']}." if row["Confidence"] else "",
            row["Important notes"],
            row["Estimate note / basis"],
        )

        evisa_row = evisa_summary.get(country_name.lower(), {})
        evisa_note = notes_text(
            evisa_row.get("eVisa / online visa availability for Ethiopian citizens", ""),
            evisa_row.get("eVisa-compatible visa types", ""),
            evisa_row.get("eVisa notes / official source", ""),
        )
        if evisa_note:
            notes = country_obj.setdefault("country_level_notes", [])
            labeled_note = f"Online visa availability note: {evisa_note}"
            if labeled_note not in notes:
                notes.append(labeled_note)

        country_obj["official_sources"] = merge_sources(
            country_obj.get("official_sources", []),
            source_entries(country_name, urls, "Workbook official source for fees, requirements and application route"),
        )

        visa_obj = find_or_create_visa_type(country_obj, visa_title, visa_slug)
        requirements = split_requirements(row["Detailed requirements"])
        steps = split_steps(row["Step-by-step guide"])
        fee = row["Government Fee / Price"] or row["Estimated fee if exact price hidden/variable"] or "Check official fee page"
        service_fee = row["Extra fees / portal notes"] or "Check official fee page"
        validity = row["Validity / Stay"] or "Verify on the official source"

        visa_obj["summary"] = (
            f"{visa_title} guide for Ethiopian passport holders applying for {country_name}, "
            f"including workbook-imported 2026 fee guidance, requirements, application route and official source links."
        )
        if requirements:
            visa_obj["requirements"] = requirements
            visa_obj["documents_needed"] = build_document_list(requirements)
        if steps:
            visa_obj["application_steps"] = steps

        visa_obj["fees"] = {
            **visa_obj.get("fees", {}),
            "visa_fee": {
                "amount": fee,
                "approx_usd": "Shown in source currency or ETB where the official source provides it.",
                "notes": notes_text(
                    "Imported from the cleaned May 2026 workbook.",
                    source_note,
                    "Always verify on the official source before payment.",
                ),
            },
            "service_fee": {
                "amount": service_fee,
                "notes": "Portal, visa-center, courier, biometrics, insurance, bank or exchange-rate costs may be separate.",
            },
        }
        visa_obj["validity_and_stay"] = {
            **visa_obj.get("validity_and_stay", {}),
            "visa_validity": validity,
            "max_stay": validity,
            "notes": "Validity and allowed stay can be different; follow the official decision notice or visa sticker.",
        }
        visa_obj["appointment_and_biometrics"] = {
            **visa_obj.get("appointment_and_biometrics", {}),
            "where_to_apply_from_ethiopia": where_to_apply or "Verify on the official source",
            "notes": notes_text("Use the official application route only.", row["Extra fees / portal notes"]),
        }
        visa_obj["application_portal_or_form"] = {
            **visa_obj.get("application_portal_or_form", {}),
            "name": where_to_apply or f"{country_name} official visa application route",
            "url": urls[0] if urls else visa_obj.get("application_portal_or_form", {}).get("url", ""),
            "title": f"Where to apply for {visa_title}",
            "notes": where_to_apply or "Verify the exact route on the official source.",
        }
        visa_obj["ethiopia_specific_notes"] = list(
            dict.fromkeys(
                [
                    *visa_obj.get("ethiopia_specific_notes", []),
                    row["Important notes"],
                    row["Estimate note / basis"],
                    evisa_note,
                ]
            )
        )
        visa_obj["official_sources"] = merge_sources(
            visa_obj.get("official_sources", []),
            source_entries(country_name, urls, f"{visa_title} fees, requirements and where to apply"),
        )
        visa_obj["data_confidence"] = row["Confidence"] or visa_obj.get("data_confidence", "Medium")
        visa_obj["missing_or_uncertain_fields"] = [
            "Fees and exchange rates can change without notice.",
            "Visa-center service fees, priority options and courier fees may vary by appointment date.",
            "Rows marked Medium or Low confidence should be rechecked before payment.",
        ]

        grouped_fee_rows.setdefault(slug, []).append(
            {
                "type": fee_type_label,
                "fee": fee,
                "what_to_prepare": "; ".join(requirements[:4]) if requirements else "Use the official checklist before applying.",
                "_first_url": urls[0] if urls else "",
                "_note": notes_text(source_note, evisa_note),
            }
        )

    for slug, rows in grouped_fee_rows.items():
        first_url = next((row["_first_url"] for row in rows if row["_first_url"]), "")
        note_bits = [row["_note"] for row in rows if row["_note"]]
        fee_guides[slug] = {
            "fee_source_url": first_url or fee_guides.get(slug, {}).get("fee_source_url", ""),
            "source_note": notes_text(
                "Imported from visa_menged_all_50_countries_estimates_cleaned.xlsx prepared May 6, 2026.",
                " ".join(dict.fromkeys(note_bits)),
                "Use these amounts for planning only and verify the official source before paying.",
            ),
            "types": [
                {key: value for key, value in row.items() if not key.startswith("_")}
                for row in rows
            ],
        }

    fee_guides = sanitize_value(fee_guides)
    research = sanitize_value(research)

    FEE_GUIDES_PATH.write_text(json.dumps(fee_guides, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    RESEARCH_PATH.write_text(json.dumps(research, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Imported {len(workbook_rows)} workbook rows into {len(grouped_fee_rows)} country fee guides.")


if __name__ == "__main__":
    main()
